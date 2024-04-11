import warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
import openpyxl as xl
import pandas as pd
from pathlib import Path
import numpy as np

filepath=Path('out.csv')

df=pd.read_excel('PCB pull.xlsx',sheet_name='PCB pull')
out=pd.DataFrame(columns=['Characteristic Name[CHANM]','Characteristic Value[CHAVL]','Target Hierarchy Name[TGT_HRY_HIENM]','Target Hierarchy Node Name[TGT_HRY_NODENAME]','Target Hierarchy Node Object Name[TGT_HRY_NODEOBJNM]','Mappable'])
unmappable=pd.DataFrame(columns=['Obj','Target'])

wb=xl.load_workbook('PyDump.xlsx')

ws=wb["Dump"]
mx=ws.max_row+1


i=2
while i<mx:

    d = ws.cell(row=i, column=1)
    obj=d.value

    id=d.value[0]

    if id == "C":
        typ="FAAKOSTL"

    elif id == "P":
        typ="FAAPOSID"
    elif id=="I":
        typ="FAACAUFN"
    else:
        typ="0HIER_NODE"


    mcloud=obj.split(" ")
    x=len(mcloud[0])
    str=mcloud[0]
    str=str.replace("$"," ")

    if typ=='0HIER_NODE':
        formapping = str[0:x]
    else:
        formapping=str[1:x]

    f = ws.cell(row=i, column=2).value
    node=f[0:10]
    targetnode=node
    objecttype=typ


    g=np.any(df['NODENAME'].isin([str]))
    if g==False:
        mapp='No'
    else:
        mapp='Yes'

    out=out._append({'Characteristic Name[CHANM]':objecttype,'Characteristic Value[CHAVL]':formapping,'Target Hierarchy Name[TGT_HRY_HIENM]':'GCOH','Target Hierarchy Node Name[TGT_HRY_NODENAME]':targetnode,'Target Hierarchy Node Object Name[TGT_HRY_NODEOBJNM]':'0HIER_NODE','Mappable':mapp},ignore_index=True)


    i+=1




out.to_csv(filepath)

